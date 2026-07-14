"""Build a per-endpoint test-coverage workbook from a test suite + endpoint list.

One Summary sheet + one sheet per endpoint (sheet name = the endpoint's section
title). Each test is attributed to the endpoint of the client method it calls.

Usage:
  python build_coverage_matrix.py <out_xlsx> <test_dirs> <endpoints_json> \
      [<extra_test_files_csv>] [<system_test_dirs_csv>]

  test_dirs           one or more comma-separated dirs (repo-relative or absolute)
  endpoints_json      JSON list of [path, section_title, method] in doc order
  extra_test_files    optional comma-separated extra test files to include
  system_test_dirs    optional comma-separated E2E/system dirs; each system test
                      is attributed to EVERY in-scope endpoint it calls, tagged
                      test-type "system" (multi-attribution)

Environment:
  COVERAGE_REPO         repo root the tests live in (default: current dir)
  COVERAGE_CLIENTS_GLOB glob for API-client modules, relative to repo
                        (default: "src/clients/*.py")

Assumptions (portable across httpx-style API-client suites):
  * client methods issue requests via ``self.api.post("/path", ...)`` /
    ``self.api.get(...)`` — the first literal (or f-string) path in a method body
    is that method's endpoint;
  * tests are collected by pytest and use allure ``@allure.title`` + markers;
  * a 4-digit case id in the title / parametrize id is treated as the Testmo id.
Tune the two ``first_path`` predicates below if a suite wraps requests differently.
"""
import sys, ast, re, os, json, glob, subprocess
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font

REPO = os.environ.get("COVERAGE_REPO") or os.getcwd()
CLIENTS_GLOB = os.environ.get("COVERAGE_CLIENTS_GLOB", "src/clients/*.py")

OUT, TEST_DIR, EP_JSON = sys.argv[1], sys.argv[2], sys.argv[3]
TDIRS = [d for d in TEST_DIR.split(",") if d]
EXTRA = sys.argv[4].split(",") if len(sys.argv) > 4 and sys.argv[4] else []
endpoints = [tuple(x) for x in json.load(open(EP_JSON))]   # (path, section, method)
EP_PATHS = {p for p, _, _ in endpoints}


# ---- 1. client method -> endpoint path (parse all client modules) ----
def first_path(fn):
    for n in ast.walk(fn):
        if isinstance(n, ast.Call) and isinstance(n.func, ast.Attribute) and n.func.attr in ("post", "get") \
           and isinstance(n.func.value, ast.Attribute) and n.func.value.attr == "api" and n.args:
            a = n.args[0]
            if isinstance(a, ast.Constant):
                return a.value
            if isinstance(a, ast.JoinedStr):   # f-string path -> keep {placeholders}
                s = ""
                for v in a.values:
                    s += v.value if isinstance(v, ast.Constant) else "{" + (v.value.id if isinstance(v.value, ast.Name) else "x") + "}"
                return s
    return None


method_to_paths = defaultdict(set)
for cf in glob.glob(os.path.join(REPO, CLIENTS_GLOB)):
    try:
        tree = ast.parse(open(cf).read())
    except Exception:
        continue
    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            p = first_path(node)
            if p:
                method_to_paths[node.name].add(p)


# ---- 2. collect test nodes ----
def _collect(dirs):
    r = subprocess.run(
        [sys.executable, "-m", "pytest", *dirs, "--collect-only", "-q", "-p", "no:cacheprovider", "-o", "addopts="],
        cwd=REPO, capture_output=True, text=True, env={**os.environ, "PYTHONPATH": REPO},
    )
    return [l.strip() for l in r.stdout.splitlines() if "::" in l]


nodes = _collect([*TDIRS, *EXTRA])


# ---- 3. AST info per test: title, marks, parametrize testmo map, called endpoints ----
def parse_parametrize(dec):
    if not (isinstance(dec, ast.Call) and getattr(dec.func, "attr", "") == "parametrize"):
        return {}
    if len(dec.args) < 2 or not isinstance(dec.args[0], ast.Constant):
        return {}
    names = [n.strip() for n in dec.args[0].value.split(",")]
    ti = next((i for i, n in enumerate(names) if n in ("testmo", "testmo_id")), -1)
    ids = None
    for kw in dec.keywords:
        if kw.arg == "ids" and isinstance(kw.value, (ast.List, ast.Tuple)):
            ids = [e.value if isinstance(e, ast.Constant) else None for e in kw.value.elts]
    pmap = {}
    if isinstance(dec.args[1], (ast.List, ast.Tuple)):
        for i, el in enumerate(dec.args[1].elts):
            tv = cid = None
            if isinstance(el, ast.Call) and getattr(el.func, "attr", "") == "param":
                if ti >= 0 and len(el.args) > ti and isinstance(el.args[ti], ast.Constant):
                    tv = el.args[ti].value
                for kw in el.keywords:
                    if kw.arg == "id" and isinstance(kw.value, ast.Constant):
                        cid = kw.value.value
            elif isinstance(el, (ast.Tuple, ast.List)) and ti >= 0 and len(el.elts) > ti and isinstance(el.elts[ti], ast.Constant):
                tv = el.elts[ti].value
            key = cid or (ids[i] if ids and i < len(ids) else None)
            if key is not None and tv is not None:
                pmap[key] = tv
    return pmap


def fn_info(fn):
    title = None
    marks = []
    pmap = {}
    for d in fn.decorator_list:
        if isinstance(d, ast.Call) and isinstance(d.func, ast.Attribute) and d.func.attr == "title" and d.args and isinstance(d.args[0], ast.Constant):
            title = d.args[0].value
        node = d.func if isinstance(d, ast.Call) else d
        if isinstance(node, ast.Attribute) and isinstance(node.value, ast.Attribute) and node.value.attr == "mark":
            marks.append(node.attr)
        pmap.update(parse_parametrize(d))
    calls = []
    for n in ast.walk(fn):
        if isinstance(n, ast.Call) and isinstance(n.func, ast.Attribute):
            # raw .api.post("/path", ...) with a literal in-scope path
            if n.func.attr in ("post", "get") and n.args and isinstance(n.args[0], ast.Constant) and n.args[0].value in EP_PATHS:
                calls.append((n.args[0].value, n.args[0].value)); continue
            paths = method_to_paths.get(n.func.attr, set()) & EP_PATHS
            if paths:
                calls.append((n.func.attr, sorted(paths)[0]))
    return title, marks, pmap, calls


def _ast_info(files):
    out = {}
    for cf in files:
        try:
            tree = ast.parse(open(cf).read())
        except Exception:
            continue
        fname = os.path.basename(cf)
        for node in tree.body:
            if isinstance(node, ast.ClassDef):
                for it in node.body:
                    if isinstance(it, (ast.FunctionDef, ast.AsyncFunctionDef)) and it.name.startswith("test"):
                        t, mk, pm, calls = fn_info(it)
                        out[(fname, node.name, it.name)] = {"title": t, "marks": mk, "pmap": pm, "calls": calls}
                        out.setdefault((fname, None, it.name), {"title": t, "marks": mk, "pmap": pm, "calls": calls})
            elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name.startswith("test"):
                t, mk, pm, calls = fn_info(node)
                out[(fname, None, node.name)] = {"title": t, "marks": mk, "pmap": pm, "calls": calls}
    return out


_absdirs = [d if os.path.isabs(d) else os.path.join(REPO, d) for d in TDIRS]
_files = [cf for d in _absdirs for cf in glob.glob(os.path.join(d, "test_*.py"))]
_files += [e if os.path.isabs(e) else os.path.join(REPO, e) for e in EXTRA]
info = _ast_info(_files)


def ttype(mk):
    for t in ("negative", "positive", "security", "smoke"):
        if t in mk:
            return t
    return ""


def resolve(title, tokens, testmo):
    if not title:
        return None
    out = title.replace("{testmo_id}", str(testmo) if testmo else "").replace("{testmo}", str(testmo) if testmo else "")
    others = [t for t in tokens if t != str(testmo)]
    for ph in re.findall(r"\{(\w+)\}", out):
        out = out.replace("{%s}" % ph, others.pop(0) if others else "", 1)
    return re.sub(r"\s{2,}", " ", out).strip().rstrip(":").strip()


# endpoint token = last path segment; plus section-title words for tie-breaks
EP_TOKENS = sorted([(p.rstrip("/").split("/")[-1], p) for p, _, _ in endpoints], key=lambda t: -len(t[0]))
_STOP = {"get", "all", "by", "to", "from", "the", "of", "and", "a", "an"}
EP_TITLEWORDS = {p: [w for w in re.findall(r"[a-z0-9]+", sec.lower()) if len(w) > 2 and w not in _STOP]
                 for p, sec, _ in endpoints}


def endpoint_for(method, calls):
    norm = method[5:] if method.startswith("test_") else method
    nwords = set(re.findall(r"[a-z0-9]+", norm.lower()))

    def name_match(paths):
        for tok, p in EP_TOKENS:
            if (norm == tok or norm.startswith(tok + "_")) and p in paths:
                return p
        return None

    def title_match(paths):
        best, bestn, tie = None, 0, False
        for p in paths:
            c = sum(1 for w in set(EP_TITLEWORDS.get(p, [])) if w in nwords)
            if c > bestn:
                best, bestn, tie = p, c, False
            elif c == bestn and c > 0 and p != best:
                tie = True
        return best if (bestn > 0 and not tie) else None

    distinct = list(dict.fromkeys(p for _, p in calls))
    if len(distinct) == 1:
        return distinct[0]
    if distinct:
        return name_match(set(distinct)) or title_match(set(distinct)) or name_match(EP_PATHS) or distinct[0]
    return name_match(EP_PATHS) or title_match(EP_PATHS)


def _node_fields(node):
    path, *rest = node.split("::")
    fname = os.path.basename(path)
    cls = rest[0] if len(rest) == 3 else (rest[0] if len(rest) == 2 and rest[0][0].isupper() else None)
    mfull = rest[-1]
    m = re.match(r"([^\[]+)(?:\[(.*)\])?$", mfull)
    return path, fname, cls, m.group(1), (m.group(2) or "")


def _testmo(tokens, param, pmap, title):
    for tok in tokens:
        if re.fullmatch(r"\d{4}", tok):
            return tok
    if param in pmap:
        return str(pmap[param])
    if title:
        mm = re.search(r"(\d{4})", title)
        return mm.group(1) if mm else ""
    return ""


groups = defaultdict(list)
unassigned = []
pending = []
file_eps = defaultdict(list)
for node in nodes:
    _, fname, cls, method, param = _node_fields(node)
    tokens = param.split("-") if param else []
    rec = info.get((fname, cls, method)) or info.get((fname, None, method)) or {}
    title, marks, pmap, calls = rec.get("title"), rec.get("marks", []), rec.get("pmap", {}), rec.get("calls", [])
    testmo = _testmo(tokens, param, pmap, title)
    name = resolve(title, tokens, testmo) or (method + (f"[{param}]" if param else ""))
    if param and not any(tok and tok in name for tok in tokens):
        name = f"{name} [{param}]"
    ep = endpoint_for(method, calls)
    row = [fname, name, ttype(marks), testmo]
    pending.append((row, ep))
    if ep:
        file_eps[fname].append(ep)
# helper/fixture-driven tests: inherit file- then service-dominant endpoint
svc_mode = Counter(ep for _, ep in pending if ep).most_common(1)
svc_ep = svc_mode[0][0] if svc_mode else None
for row, ep in pending:
    if not ep:
        fmode = Counter(file_eps.get(row[0], [])).most_common(1)
        ep = fmode[0][0] if fmode else svc_ep
    if ep:
        groups[ep].append(row)
    else:
        unassigned.append(row)

# ---- 3b. system/E2E tests: attribute to EVERY in-scope endpoint they call ----
SYS = sys.argv[5].split(",") if len(sys.argv) > 5 and sys.argv[5] else []
if SYS:
    sysabs = [d if os.path.isabs(d) else os.path.join(REPO, d) for d in SYS]
    sysfiles = [p for d in sysabs for p in glob.glob(os.path.join(d, "**", "test_*.py"), recursive=True)]
    sysinfo = _ast_info(sysfiles)
    for node in _collect(sysabs):
        path, fname, cls, method, param = _node_fields(node)
        tokens = param.split("-") if param else []
        rec = sysinfo.get((fname, cls, method)) or sysinfo.get((fname, None, method)) or {}
        title, pmap, calls = rec.get("title"), rec.get("pmap", {}), rec.get("calls", [])
        testmo = _testmo(tokens, param, pmap, title)
        name = resolve(title, tokens, testmo) or method
        if param and not any(tok and tok in name for tok in tokens):
            name = f"{name} [{param}]"
        rel = os.path.relpath(os.path.join(REPO, path), os.path.join(REPO, "tests")) if not os.path.isabs(path) else path
        for p in dict.fromkeys(pp for _, pp in calls):
            groups[p].append([rel, name, "system", testmo])

# ---- 4. write workbook ----
wb = openpyxl.Workbook()
wb.remove(wb.active)
BOLD = Font(bold=True)
sm = wb.create_sheet("Summary")
sm.append(["Endpoint", "Method", "Path", "# tests", "positive", "negative", "other"])
for c in sm[1]:
    c.font = BOLD
total = tp = tn = 0
for p, sec, meth in endpoints:
    rows = groups.get(p, [])
    pos = sum(1 for x in rows if x[2] == "positive")
    neg = sum(1 for x in rows if x[2] == "negative")
    sm.append([sec, meth, p, len(rows), pos, neg, len(rows) - pos - neg])
    total += len(rows); tp += pos; tn += neg
if unassigned:
    sm.append(["(unassigned)", "", "", len(unassigned), "", "", ""])
sm.append([])
sm.append(["TOTAL", "", "", total + len(unassigned), tp, tn, ""])
for c in sm[sm.max_row]:
    c.font = BOLD
for col, w in zip("ABCDEFG", [26, 8, 40, 8, 9, 9, 7]):
    sm.column_dimensions[col].width = w
for p, sec, meth in endpoints:
    s = wb.create_sheet(sec[:31])
    s["A1"] = f"{meth} {p}"; s["A1"].font = BOLD
    s.append(["File name", "Test name", "Test type", "Testmo ID"])
    for c in s[2]:
        c.font = BOLD
    for r in sorted(groups.get(p, []), key=lambda x: (x[3] or "zzzz")):
        s.append(r)
    for col, w in zip("ABCD", [28, 66, 10, 10]):
        s.column_dimensions[col].width = w
    s.freeze_panes = "A3"
if unassigned:
    s = wb.create_sheet("(unassigned)")
    s.append(["File name", "Test name", "Test type", "Testmo ID"])
    for c in s[1]:
        c.font = BOLD
    for r in unassigned:
        s.append(r)
wb.save(OUT)
print("nodes:", len(nodes), "assigned:", total, "unassigned:", len(unassigned))
for p, sec, meth in endpoints:
    print(f"  {len(groups.get(p, [])):3d}  {sec}")
if unassigned:
    print("UNASSIGNED:")
    for u in unassigned:
        print("   ", u[0], "|", u[1][:55])
