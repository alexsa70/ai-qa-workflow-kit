#!/usr/bin/env python3
"""Render a Test Design Contract (.md, kit template) into a readable .xlsx.

The .md contract is the source of truth; this writes a derived, read-friendly
spreadsheet next to it (same base name, .xlsx). One row per case on a
`Test Cases` sheet, one row per step on a `Steps` sheet, plus a `Contract`
metadata sheet.

Usage:
    python contract_to_xlsx.py path/to/contract.md [-o path/to/out.xlsx]

Requires openpyxl:  pip install openpyxl --break-system-packages
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl --break-system-packages")

CASE_COLUMNS = [
    "ID", "Name", "Priority", "Request actor", "Target variants",
    "Preconditions", "Cleanup", "Evidence",
]

STEP_COLUMNS = [
    "Case ID", "Step #", "Action", "Request", "Example payload",
    "Status", "Message", "Value checks", "Schema", "Headers",
    "Error contract", "Side effects",
]

# Map contract bullet labels -> case dict keys (case-level fields only; step
# detail lives under the '#### Steps' section and is parsed separately).
TOP_LABELS = {
    "priority": "Priority",
    "request actor": "Request actor",
    "target variants": "Target variants",
    "given": "Preconditions",
    "cleanup": "Cleanup",
    "evidence": "Evidence",
}

# Step-level top bullets (directly under a numbered step).
STEP_TOP_LABELS = {
    "request": "Request",
    "example payload": "Example payload",
}

# Bullets nested under a step's '- Pass criteria:' line.
PASS_LABELS = {
    "status": "Status",
    "message": "Message",
    "value checks": "Value checks",
    "schema": "Schema",
    "headers": "Headers",
    "error contract": "Error contract",
    "side effects": "Side effects",
}

META_LABELS = ["Project", "Status", "Requirement source", "Scope",
               "Exclusions", "Next transition"]


def parse_contract(text: str):
    meta = {}
    for label in META_LABELS:
        m = re.search(rf"^- {re.escape(label)}:\s*(.*)$", text, re.MULTILINE | re.IGNORECASE)
        if m:
            meta[label] = m.group(1).strip().strip("`")

    cases = []
    all_steps = []
    cases_section = _extract_section(text, "Cases")
    # Split on case headers: ### TC-001: name
    parts = re.split(r"^###\s+(TC-[^\s:]+):?\s*(.*)$", cases_section, flags=re.MULTILINE)
    # parts: [pre, id1, name1, body1, id2, name2, body2, ...]
    for i in range(1, len(parts), 3):
        cid = parts[i].strip()
        name = parts[i + 1].strip()
        body = parts[i + 2]

        steps_text, rest_body = _extract_steps_section(body)

        case = {c: "" for c in CASE_COLUMNS}
        case["ID"] = cid
        case["Name"] = name
        _parse_case_top_fields(rest_body, case)
        cases.append(case)

        all_steps.extend(_parse_steps(steps_text, cid))

    return meta, cases, all_steps


def _extract_section(text: str, heading: str) -> str:
    """Return the body of a top-level '## <heading>' section, stopping at the
    next '## ' heading or end of document."""
    m = re.search(rf"^##\s+{re.escape(heading)}\s*$", text, re.MULTILINE)
    if not m:
        return ""
    start = m.end()
    end_m = re.search(r"^##\s+\S", text[start:], re.MULTILINE)
    end = start + end_m.start() if end_m else len(text)
    return text[start:end]


def _extract_steps_section(body: str):
    """Split a case body into (steps_section_text, body_without_steps)."""
    m = re.search(r"^####\s+Steps\s*$", body, re.MULTILINE)
    if not m:
        return "", body
    start = m.end()
    end_match = re.search(r"^(####\s|- (Cleanup|Evidence):)", body[start:], re.MULTILINE)
    end = start + end_match.start() if end_match else len(body)
    steps_text = body[start:end]
    rest_body = body[:m.start()] + body[end:]
    return steps_text, rest_body


def _parse_case_top_fields(body: str, case: dict):
    lines = body.splitlines()
    cur_key = None
    buf: list[str] = []

    def flush():
        nonlocal buf, cur_key
        if cur_key and buf:
            val = "\n".join(b.rstrip() for b in buf).strip()
            case[cur_key] = (case[cur_key] + "\n" + val).strip() if case[cur_key] else val
        buf = []

    for raw in lines:
        line = raw.rstrip()
        m_top = re.match(r"^- ([A-Za-z][A-Za-z /]+):\s*(.*)$", line)
        if m_top:
            flush()
            label = m_top.group(1).strip().lower()
            rest = m_top.group(2).strip()
            cur_key = TOP_LABELS.get(label)
            buf = [rest] if rest else []
        else:
            stripped = line.strip()
            if stripped.startswith("- "):
                stripped = stripped[2:]
            if stripped and cur_key:
                buf.append(stripped)
    flush()


def _parse_steps(text: str, cid: str) -> list[dict]:
    steps = []
    parts = re.split(r"^(\d+)\.\s*(.*)$", text, flags=re.MULTILINE)
    # parts: [pre, num1, action1, body1, num2, action2, body2, ...]
    for i in range(1, len(parts), 3):
        num = parts[i].strip()
        action = parts[i + 1].strip()
        body = parts[i + 2]
        step = {c: "" for c in STEP_COLUMNS}
        step["Case ID"] = cid
        step["Step #"] = num
        step["Action"] = action
        _parse_step_body(body, step)
        steps.append(step)
    return steps


def _parse_step_body(body: str, step: dict):
    lines = body.splitlines()
    cur_key = None
    in_pass = False
    pass_indent = None
    buf: list[str] = []

    def leading(s: str) -> int:
        return len(s) - len(s.lstrip(" "))

    def flush():
        nonlocal buf, cur_key
        if cur_key and buf:
            val = "\n".join(b.rstrip() for b in buf).strip()
            step[cur_key] = (step[cur_key] + "\n" + val).strip() if step[cur_key] else val
        buf = []

    for raw in lines:
        line = raw.rstrip()
        if not line.strip():
            continue
        indent = leading(line)
        m = re.match(r"^\s*- ([A-Za-z][A-Za-z /]+):\s*(.*)$", line)
        if m:
            label = m.group(1).strip().lower()
            rest = m.group(2).strip()
            if in_pass and pass_indent is not None and indent > pass_indent:
                flush()
                cur_key = PASS_LABELS.get(label)
                buf = [rest] if rest else []
                continue
            flush()
            if label == "pass criteria":
                in_pass = True
                pass_indent = indent
                cur_key = None
                continue
            in_pass = False
            cur_key = STEP_TOP_LABELS.get(label)
            buf = [rest] if rest else []
        else:
            stripped = line.strip()
            if stripped.startswith("- "):
                stripped = stripped[2:]
            if stripped and cur_key:
                buf.append(stripped)
    flush()


def build_workbook(meta: dict, cases: list[dict], steps: list[dict]) -> Workbook:
    wb = Workbook()
    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Test Cases"
    ws.append(CASE_COLUMNS)
    for c in ws[1]:
        c.font = header_font
        c.alignment = Alignment(vertical="top")
    for case in cases:
        ws.append([case.get(col, "") for col in CASE_COLUMNS])

    case_widths = {
        "ID": 10, "Name": 40, "Priority": 10, "Request actor": 16,
        "Target variants": 20, "Preconditions": 40, "Cleanup": 25,
        "Evidence": 20,
    }
    for idx, col in enumerate(CASE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = case_widths.get(col, 20)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_top
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CASE_COLUMNS))}1"

    ws_steps = wb.create_sheet("Steps")
    ws_steps.append(STEP_COLUMNS)
    for c in ws_steps[1]:
        c.font = header_font
        c.alignment = Alignment(vertical="top")
    for step in steps:
        ws_steps.append([step.get(col, "") for col in STEP_COLUMNS])

    step_widths = {
        "Case ID": 10, "Step #": 8, "Action": 35, "Request": 22,
        "Example payload": 30, "Status": 10, "Message": 25,
        "Value checks": 30, "Schema": 18, "Headers": 18,
        "Error contract": 22, "Side effects": 22,
    }
    for idx, col in enumerate(STEP_COLUMNS, start=1):
        ws_steps.column_dimensions[get_column_letter(idx)].width = step_widths.get(col, 18)
    for row in ws_steps.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_top
    ws_steps.freeze_panes = "A2"
    ws_steps.auto_filter.ref = f"A1:{get_column_letter(len(STEP_COLUMNS))}1"

    ws2 = wb.create_sheet("Contract")
    ws2.append(["Field", "Value"])
    for c in ws2[1]:
        c.font = header_font
    for label in META_LABELS:
        ws2.append([label, meta.get(label, "")])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70
    for row in ws2.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    return wb


def main():
    ap = argparse.ArgumentParser(description="Render a test design contract .md to a readable .xlsx.")
    ap.add_argument("contract", help="path to the contract .md")
    ap.add_argument("-o", "--out", help="output .xlsx (default: contract path with .xlsx)")
    args = ap.parse_args()

    src = Path(args.contract)
    if not src.is_file():
        sys.exit(f"contract not found: {src}")
    out = Path(args.out) if args.out else src.with_suffix(".xlsx")

    meta, cases, steps = parse_contract(src.read_text(encoding="utf-8"))
    if not cases:
        print("warning: no TC-* cases found; writing metadata-only workbook", file=sys.stderr)
    wb = build_workbook(meta, cases, steps)
    wb.save(out)
    print(f"wrote {len(cases)} cases, {len(steps)} steps -> {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
